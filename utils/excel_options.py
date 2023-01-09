import openpyxl
from openpyxl.styles import Font, PatternFill, NamedStyle
from string import ascii_uppercase


def change_style(workbook):
    """Creates a default named style to be used in the Excel file."""
    default = NamedStyle(name="default")
    default.font = Font(name='Bahnschrift', color='FFFFFF')
    default.fill = PatternFill(patternType='solid', start_color='222222', end_color='222222')
    workbook.add_named_style(default)
    return default


def create_calendar_sheet(workbook, style, calendar, tournaments):
    """
    Creates a sheet in the given workbook with the World Cup calendar data.
    Parameters:
        workbook (openpyxl.Workbook): Workbook object to add the sheet to.
        style (openpyxl.styles.NamedStyle): Style to apply to the cells in the sheet.
        calendar (list): List of lists containing the calendar data. Each inner list should contain
            the competition ID, hill name, and status.
        tournaments (list): List of lists containing the data for each tournament. Each inner list
            should contain the tournament name, competition type, and a string of comma-separated
            competition IDs.
    """
    # Prepare tournaments for adding to the sheet
    tournament_data = []
    tournament_names = []
    # Iterate through the tournaments and create a list of tournament names
    # and a list of tuples containing the competition IDs and the column letter
    for tournament, col in zip(tournaments, ascii_uppercase[3:]):
        name, *_, comp_ids = tournament
        comp_ids = [int(comp_id)+2 for comp_id in comp_ids.split(',')]
        tournament_names.append(name)
        tournament_data.append([comp_ids, col])

    # Add the data to the sheet
    sheet = workbook.create_sheet('WC Calendar')
    sheet.append(['World Cup Calendar'])
    sheet.append(['Id', 'Hill', 'Status', *tournament_names])
    for competition in calendar:
        sheet.append(competition)
    # Add an 'X' to the cell in the column for each tournament if the competition
    # is part of the tournament
    for rows, column in tournament_data:
        for cell in sheet[column]:
            if cell.row in rows:
                sheet[cell.coordinate] = 'X'

    # Change the style for all cells in the sheet
    for col in ascii_uppercase:
        for cell in sheet[col]:
            cell.style = style
    # Change the column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 15
    for col in ascii_uppercase[3:]:
        sheet.column_dimensions[col].width = 20


def create_classification_sheets(workbook, style, classifications):
    """Creates sheets in the given workbook with the classification data.
     Parameters:
        workbook (openpyxl.Workbook): Workbook object to add the sheets to.
        style (openpyxl.styles.NamedStyle): Style to apply to the cells in the sheets.
        classifications (list): List of tuples containing the classification data."""
    for i, classification in enumerate(classifications):
        name, results = classification
        sheet = workbook.create_sheet(name[:30])
        sheet.append([name])
        for row in results:
            sheet.append(row)
        # Change theme for affected rows
        for col in ascii_uppercase:
            sheet.column_dimensions[col].width = 10
            for cell in sheet[col]:
                cell.style = style
        # Competitor name column wider
        sheet.column_dimensions['B'].width = 30


def save_results_in_excel(classifications, calendar, tournaments):
    """Saves the classifications in multiple sheets of Excel file."""
    # Create Workbook object
    workbook = openpyxl.Workbook()
    default = change_style(workbook)
    # Create first sheet with calendar
    create_calendar_sheet(workbook, default, calendar, tournaments)
    # Write every classification to different sheet
    create_classification_sheets(workbook, default, classifications)
    # Remove default sheet
    workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
    # Save workbook
    workbook.save('./data/classifications.xlsx')
